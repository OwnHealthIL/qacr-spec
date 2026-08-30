#!/usr/bin/env python3
"""Renders Generator/m5-map.js as a workbook for Guy to mark up.

    python3 Generator/build-m5-proposal.py

One-off, like spec-triage.py. Nothing depends on its output; the implementation reads
m5-map.js directly, so approving the proposal and applying it use the same map.
"""
import json, os, subprocess, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(DIR)
OUT = os.path.join(ROOT, "QACR-APP-FR-01 Milestone 5 proposal.xlsx")

def node(expr):
    return json.loads(subprocess.check_output(
        ["node", "-e", f"console.log(JSON.stringify({expr}))"], cwd=DIR).decode())

MAP = node("require('./m5-map.js')")
BACKLOG = node("require('./appendices.js').backlog")
EPICS = node("require('./epics.js')")
SECTIONS = node("[...require('./reqs-part1.js'), ...require('./reqs-part2.js')]")

TITLE = {}
for s in SECTIONS:
    TITLE[s["reqs"][0][0].split("-")[1]] = s["title"]

BL = {r[0]: r for r in BACKLOG}
OWNER = {}
for e in EPICS:
    for f in e["features"]:
        for i in f[3]:
            OWNER[i] = (f[0], f[1], e["code"])

ARIAL = "Arial"
HEAD = PatternFill("solid", fgColor="1F3864")
BAND = PatternFill("solid", fgColor="F2F5FA")
FLAG = PatternFill("solid", fgColor="FFF3CD")
INPUT = PatternFill("solid", fgColor="FFFFCC")
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ----------------------------------------------------------------- Read me
ws = wb.active
ws.title = "Read me"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 112
rows = [
    ("h", "Milestone 5 — pulling the backlog into the requirements"),
    ("", ""),
    ("p", "You asked to empty the backlog, put every item into the section it belongs to, and give them all a "
          "new milestone 5, described as future development, with no date. This is that proposal. Nothing has "
          "been changed yet."),
    ("", ""),
    ("h2", "The three decisions you already made"),
    ("p", "1. Items that were once real requirements go back to their original identifier. BL-25 becomes "
          "FR-ACC-003 again, unchanged text, now milestone 5. Fourteen items qualify and every one of those "
          "numbers is still vacant, so nothing collides."),
    ("p", "2. The four post-test survey items go into FR-ANL, which is already titled Analytics and User "
          "Feedback. FR-SUR is not reinstated, so this morning's prefix rule and its guard stand."),
    ("p", "3. Milestone 5 requirements sit in the body, in their own sections. The document goes from 199 "
          "requirements to 241."),
    ("", ""),
    ("h2", "What I need from you"),
    ("p", "The Proposal tab, one row per backlog item. Check the section each one lands in. Five rows are "
          "shaded amber because the group is genuinely arguable and I have written the alternative in the "
          "Arguable column — those are the ones worth your attention first."),
    ("p", "Put anything you want changed in the last column. Leave a row blank to accept it."),
    ("", ""),
    ("h2", "Two things I decided that you should overrule if you disagree"),
    ("p", "Existing gaps are not filled. FR-SHR-009, FR-PRT-006, FR-AUT-009 and the rest were left by "
          "withdrawn or superseded requirements and stay vacant, because an identifier is never reused for "
          "different content. Only the fourteen restorations reoccupy their own former numbers."),
    ("p", "The internal backlog priority is dropped. Every item had a 1-to-5 ordering used only to sequence "
          "the backlog. Once everything is milestone 5 that ordering has no meaning in the document. If you "
          "want an order within milestone 5, say so and it becomes a field of its own rather than a "
          "milestone."),
    ("", ""),
    ("h2", "Effects"),
    ("p", "The Effects tab lists every place a BL identifier appears today, so you can see the change is "
          "traced rather than trusted. The identifier guard will require a disposition for all 42, which is "
          "the mechanism that stops one being lost."),
]
r = 1
for kind, text in rows:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "h":
        c.font = Font(name=ARIAL, size=15, bold=True, color="1F3864")
    elif kind == "h2":
        c.font = Font(name=ARIAL, size=11, bold=True, color="1F3864")
    else:
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 15 + 12 * (len(text) // 105)
    r += 1

# ----------------------------------------------------------------- Proposal
ws = wb.create_sheet("Proposal")
cols = [("Backlog", 10), ("Proposed ID", 14), ("Section", 34), ("Disposition", 12),
        ("Feature", 9), ("Requirement", 72), ("Why this group", 52), ("Arguable — the alternative", 46),
        ("YOUR COMMENT", 30)]
for i, (h, w) in enumerate(cols, start=1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    c.fill = HEAD
    c.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# restorations first, then the rest grouped by section
order = sorted(MAP, key=lambda r: (r[2] != "restore", r[1].split("-")[1], int(r[1].split("-")[2])))
r = 2
for bl, fid, why, arguable in order:
    src = BL[bl]
    prefix = fid.split("-")[1]
    restored = why == "restore"
    owner = OWNER.get(bl, ("—", "", ""))
    vals = [
        bl, fid, f"FR-{prefix}  {TITLE.get(prefix,'')}",
        "restored" if restored else "new number",
        owner[0],
        src[1],
        "Superseded into the backlog from this exact identifier, vacant ever since. Text unchanged."
        if restored else why,
        arguable or "",
        "",
    ]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=ARIAL, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BOX
        if arguable:
            c.fill = FLAG
        elif r % 2 == 0:
            c.fill = BAND
        if i == 2:
            c.font = Font(name=ARIAL, size=9, bold=True)
        if i == 9:
            c.fill = INPUT
    ws.row_dimensions[r].height = max(28, 11 * (len(src[1]) // 70 + 1))
    r += 1

# ----------------------------------------------------------------- Effects
ws = wb.create_sheet("Effects")
for i, (h, w) in enumerate([("Where", 40), ("BL mentions", 12), ("What has to change", 84)], start=1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    c.fill = HEAD

def count(path):
    """BL mentions in one file. Never split the path on ' and ' — both spec filenames
    contain it ('Readiness and eligibility', 'Timed waits and the reaction phase')."""
    # Generated artefacts are excluded: they are rebuilt from the data, so a mention
    # count in them measures nothing a person has to change by hand.
    DERIVED = ("id-manifest.json", "Board.html", ".xlsx", ".docx")
    full = os.path.join(ROOT, path)
    if path.endswith(DERIVED):
        return "regenerated"
    if not os.path.isfile(full) or not path.endswith((".js", ".py", ".md")):
        return 0
    return len(re.findall(r"BL-\d\d", open(full, encoding="utf-8").read()))

EFFECTS = [
    ("Generator/appendices.js", "backlog array empties. The 42 items move into reqs-part1.js and reqs-part2.js as milestone-5 requirements. withdrawn is untouched."),
    ("Generator/reqs-part1.js", "Receives the milestone-5 requirements for its sections, each with a note recording where it came from."),
    ("Generator/reqs-part2.js", "The same for its sections."),
    ("Generator/epics.js", "Every feature's second identifier list held its BL items. Those move into the first list, so each feature's requirements are one set again."),
    ("Generator/review.js", "The register cites BL identifiers in the consequences-of-deferral entries and in several scope questions. Each becomes the new FR identifier. The consequences themselves remain true: milestone 5 is still not built for the submission."),
    ("Generator/configs.js", "The configuration register maps values to the requirements they realise, including BL items. Each remaps."),
    ("Generator/spta.js", "One threat-analysis mapping cites a BL item."),
    ("Generator/id-guard.js", "Learns a restored disposition, and requires all 42 BL identifiers to be dispositioned rather than simply gone."),
    ("Generator/id-manifest.json", "Regenerated. 42 BL identifiers retire, 42 FR identifiers appear, 284 on record becomes 326."),
    ("Generator/build.js", "The milestone table gains milestone 5, future development, no date. The backlog appendix goes. The review-register summary loses its deferred count."),
    ("Generator/build-epics.js", "Milestone 5 badges, and the deferred concept in feature rows and milestone spans is replaced."),
    ("Generator/build-board.js", "The Deferred filter chip and the DEF badge become milestone 5. The bl data set goes."),
    ("Generator/spec-status.js", "Seven features are described as deferred backlog items that will still be spec'd. They are now milestone 5, so the wording changes but the fact does not."),
    ("Generator/consistency-check.py", "The milestone list gains 5. Deferred handling is replaced. The count of guards rises."),
    ("Generator/spec-check.py", "Accepts BL identifiers as valid spec traces. After this only FR identifiers exist."),
    ("Specs/QACR-APP-SPEC-01 Readiness and eligibility.md", "Cites BL-07, BL-15, BL-33, BL-34, BL-35 in its configuration section, its proposals and its open items."),
    ("Specs/QACR-APP-SPEC-05 Timed waits and the reaction phase.md", "Cites BL-24 for the hardened time source and BL-33."),
    ("CLAUDE.md", "The identifier scheme drops BL-nn. The milestone table gains 5. The spec-status wording changes. Section 10's BL-34, BL-35 and BL-40 threads are restated against the new identifiers."),
    ("QACR-APP-EPIC-01 Board.html", "Rebuilt, and the zip with it. The zip you sent shows Deferred; a developer opening the new one sees milestone 5. Worth resending."),
    ("QACR-APP-SPEC-00 Spec Triage.xlsx", "Its milestone column shows an em dash for deferred features. Regenerating it is optional; the answers live in spec-status.js now."),
]
r = 2
for where, what in EFFECTS:
    n = count(where)
    for i, v in enumerate([where, n or "", what], start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=ARIAL, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BOX
        if r % 2 == 0:
            c.fill = BAND
    ws.row_dimensions[r].height = max(26, 11 * (len(what) // 82 + 1))
    r += 1

ws.cell(row=r + 1, column=1, value="Total BL mentions in hand-maintained sources").font = Font(name=ARIAL, size=9, bold=True)
ws.cell(row=r + 1, column=2, value=sum(n for n in (count(w) for w, _ in EFFECTS) if isinstance(n, int))).font = Font(name=ARIAL, size=9, bold=True)

wb.save(OUT)
print(f"wrote {OUT}")
print(f"  {len(MAP)} items proposed · {sum(1 for r in MAP if r[2]=='restore')} restored · "
      f"{sum(1 for r in MAP if r[3])} flagged as arguable")
