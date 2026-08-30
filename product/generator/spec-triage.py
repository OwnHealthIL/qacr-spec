#!/usr/bin/env python3
"""Builds the spec triage worksheet: every feature classified against Minuteful
Kidney, so the size of the spec job is known before any spec is written."""
import sys, json, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = sys.argv[1]
FEATURES = json.load(open(sys.argv[2]))

INK, F = "1F3864", "Arial"
HDRF = PatternFill("solid", fgColor=INK)
ALT = PatternFill("solid", fgColor="F2F5FA")
FILLIN = PatternFill("solid", fgColor="FFF6D6")
THIN = Side(style="thin", color="BFC9DA")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---- the triage. [status, confidence, why, what the spec needs]
# status: Unchanged | Changed | New | No spec | Deferred
T = {
 # E01 -----------------------------------------------------------------
 "F01.1": ("Unchanged", "High", "Store publication and the store-side compatibility filters work the same way; only the listing content differs.", "One paragraph: as Minuteful, plus the QACR supported-device statement."),
 "F01.2": ("Unchanged", "High", "Same mechanism and the same configuration keys — minOsVersion, minAppVersion, supportedHardwareConfig.", "One paragraph plus the three blocking alerts, which exist already."),
 "F01.3": ("Unchanged", "High", "Root and jailbreak detection, unchanged.", "One paragraph: as Minuteful."),
 "F01.4": ("Unchanged", "High", "You confirmed the pattern at review: silent while on the home screen, alerts on failure when start test is tapped.", "One paragraph, plus which check maps to which alert."),
 "F01.5": ("Unchanged", "High", "You confirmed cold-start conditions are block screens and in-flow conditions are alerts, as in Minuteful.", "The alert template and the list of conditions using it. Screenshots you already sent cover the four block screens."),
 "F01.6": ("Unchanged", "Medium", "Retransmission of an unsent test on launch. Assumed to work as Minuteful; I have not seen it.", "Confirm it exists in Minuteful and behaves the same."),
 "F01.7": ("No spec", "High", "A constraint on every screen rather than a feature. Nothing to specify beyond the requirement.", "None."),
 "F01.8": ("Unchanged", "High", "Model, OS and app version on the test record.", "One line: as Minuteful, plus any QACR field additions."),
 "F01.9": ("Changed", "High", "Broadened at review to all server configuration, and the QACR configuration set is a new set rather than the Minuteful one.", "The retrieval and validation behaviour, the failure alert, and the QACR key list once the new set exists."),
 # E02 -----------------------------------------------------------------
 "F02.1": ("Unchanged", "High", "Phone entry, format validation, backend lookup.", "One paragraph. Screenshots needed for country-code handling."),
 "F02.2": ("Unchanged", "High", "Your screenshot shows the six-digit code, resend and change-number, and you confirmed resend spends an attempt with a thirty-second interval.", "The screen, the three controls, the lockout state. Screenshot already provided."),
 "F02.3": ("Unchanged", "Medium", "You said start from something similar to Minuteful. I have not seen the invite-code screen.", "Screenshot, and where the user gets the code."),
 "F02.4": ("Changed", "High", "Date of birth is no longer at login; it is confirmed at test start, behind a feature flag for the clinical study.", "The test-start confirmation screen, the flag behaviour, the generic error."),
 "F02.5": ("Changed", "Medium", "Token lifecycle is the same mechanism, but the twenty-four-hour expiry is disputed as Q-63.", "One paragraph, and the expiry value once Q-63 is settled."),
 "F02.6": ("Unchanged", "High", "Guidance and support escalation on repeated failure.", "One paragraph: as Minuteful."),
 "F02.7": ("Unchanged", "High", "Checkbox on phone entry, policy links. Recording stays deferred as BL-03.", "One paragraph plus the checkbox states."),
 "F02.8": ("New", "Low", "Onboarding screens ahead of authentication. You added the requirement; I do not know whether Minuteful has an equivalent.", "Tell me whether Minuteful has these. If it does this becomes Unchanged."),
 # E03 -----------------------------------------------------------------
 "F03.1": ("Changed", "High", "Same QR mechanism, but the QACR code sits on the Test Board and the identifier template differs.", "The scan-time validation, the missing or damaged code path, the backend check."),
 "F03.2": ("Unchanged", "High", "The twenty-four-hour new-kit alert and scan-time reuse detection both exist in Minuteful.", "One paragraph plus the alert and its confirmation."),
 # E04 -----------------------------------------------------------------
 "F04.1": ("Changed", "High", "The chat-style sequencer exists, but the QACR step list is entirely different and FR-FLW-006 adds conditional steps.", "The sequencer behaviour, the confirmation control, and the conditional-step mechanism. The step list itself belongs to the copy file."),
 "F04.2": ("Changed", "High", "Media and replay work as Minuteful; every video and animation is new, and which steps are replayable is per step.", "The replay rules and the per-step designation. Media production is separate."),
 "F04.3": ("Unchanged", "High", "You confirmed it behaves as in Minuteful.", "One paragraph: as Minuteful."),
 "F04.4": ("Unchanged", "High", "You confirmed: start from Minuteful and improve on it.", "One paragraph, plus whatever the improvement turns out to be."),
 "F04.5": ("Changed", "High", "Cancellation is permitted only before the cup is filled, which is a QACR-specific gate.", "The two behaviours either side of cup fill, and where the exit control lives."),
 "F04.6": ("New", "High", "Demonstration navigation does not exist in Minuteful.", "Full spec, but small: the gesture, what it reveals, the run-time gate on demonstration mode."),
 "F04.7": ("New", "High", "Scan practice does not exist in Minuteful.", "Full spec: entry point, what differs from a real scan, how it ends, how the user knows it was practice."),
 "F04.8": ("New", "High", "The waiting-time card is new, and comes out of the reaction-phase research.", "Full spec: when the flow opens it, the timing thresholds, the countdown, how content rotates."),
 # E05 -----------------------------------------------------------------
 "F05.1": ("New", "Low", "QACR has two timed windows with a user action between them. Minuteful's strip test has at most one develop-and-scan wait, so I expect the framework to be new — but this is the assumption I am least sure of.", "Full spec, and the most important one in the set. Tell me what timing Minuteful actually has."),
 "F05.2": ("New", "High", "Sample-collection to Sample Pod window has no Minuteful equivalent.", "Full spec: silent timing, the block on expiry, no countdown shown."),
 "F05.3": ("New", "High", "Incubation residence time and the coloured release button do not exist in Minuteful.", "Full spec. The critical interaction of the whole test: wait, then act promptly."),
 "F05.4": ("New", "Medium", "Colour evolution then a scanning window. Minuteful has a develop-then-scan pattern, so parts may carry over.", "Full spec, informed by whatever Minuteful does for its develop window."),
 "F05.5": ("New", "Medium", "Two countdowns plus a background notification. Minuteful may show one countdown.", "Full spec: what is shown, what is silent, the local notification."),
 "F05.6": ("New", "Medium", "Invalidation on timing grounds. Minuteful invalidates on its own timing rules.", "Full spec: which timing failures invalidate, and what the user is told."),
 # E06 -----------------------------------------------------------------
 "F06.1": ("Changed", "High", "Frame-set acquisition exists; the QACR frame recipe, torch states and adjustments come from the QACR IVTS specification.", "The acquisition behaviour. The recipe itself belongs to the IVTS document, referenced not restated."),
 "F06.2": ("Changed", "High", "The QCR stage exists in Minuteful; thresholds and markers are QACR-specific.", "The stage behaviour and the accept-or-reject loop."),
 "F06.3": ("Changed", "High", "You said follow the Minuteful design. The gate set differs and specularity is new.", "The gates the Risk Analysis relies on, each with its user-facing state. Screenshots of the Minuteful treatments would settle most of it."),
 "F06.4": ("Changed", "High", "Real-time guidance exists; the spoken channel and the QACR conditions are new or changed.", "The guidance states and the spoken behaviour. Voice script is copy."),
 "F06.5": ("Changed", "Medium", "Normalisation exists in some form; whether it runs on device or in the backend is open as Q-24.", "Short spec, blocked on Q-24."),
 "F06.6": ("Changed", "High", "Post-capture boundary evaluation exists; specularity is new, and the analysis indication needs stating.", "The pipeline hand-off, the boundary outcomes, the analysis state, the retry."),
 # E07 -----------------------------------------------------------------
 "F07.1": ("No spec", "High", "Becomes the copy file's governance: versioning, reading level, approved names.", "None. Governed by the copy file."),
 "F07.2": ("No spec", "High", "Thirteen instructional strings. All copy.", "None. Goes in the copy file."),
 "F07.3": ("No spec", "High", "Result and invalid-test copy.", "None. Goes in the copy file."),
 "F07.4": ("No spec", "High", "Blocking-message content rules.", "None. Rules live with the copy file; the alert template is F01.5."),
 # E08 -----------------------------------------------------------------
 "F08.1": ("New", "High", "A different assay measured a different way. Nothing carries over from the strip algorithm.", "Owned by the algorithm team, not this spec set. Reference their document."),
 "F08.2": ("Unchanged", "High", "Version governance mechanism is the same.", "One paragraph, backend-owned."),
 "F08.3": ("New", "High", "Standard curves, recovery and blank wells, incomplete-run detection — all QACR.", "Owned by the algorithm team. Reference their document."),
 "F08.4": ("Changed", "High", "A payload contract exists; the invalid-reason enumeration is new and has no owner yet (Q-45).", "The contract, jointly with backend. Blocked on Q-45."),
 # E09 -----------------------------------------------------------------
 "F09.1": ("Changed", "High", "Minuteful shows an ACR value and category. QACR embeds the same display in the chat and in the results center.", "The display in both contexts, the categories and their treatment."),
 "F09.2": ("Changed", "Medium", "Out-of-range and invalid outcomes exist; the QACR reason enumeration is new.", "The three outcomes. Blocked in part on Q-45."),
 "F09.4": ("New", "High", "Demonstration result does not exist in Minuteful.", "Small spec: the fixed payload, the marking, where it appears."),
 # E10 -----------------------------------------------------------------
 "F10.1": ("Unchanged", "High", "Four-digit PIN, five attempts, twenty-four-hour setup window — all as Minuteful and enforced server-side.", "One paragraph plus the PIN screens. Screenshots needed."),
 "F10.2": ("Unchanged", "High", "Fifteen-minute inactivity, PIN on re-entry.", "One line: as Minuteful."),
 "F10.3": ("Changed", "High", "Split at review: main screen shows the most recent result, history deferred to BL-42.", "The main screen. Note what was removed relative to Minuteful, which shows history."),
 "F10.4": ("Unchanged", "High", "Hamburger menu, About, policies.", "One paragraph plus the menu item list."),
 "F10.5": ("Unchanged", "High", "Telephone and email support.", "One line: as Minuteful."),
 "F10.6": ("Deferred", "High", "Post-test lobby is deferred, and you already have the lobby 4 spec for it.", "None now. The existing spec is the starting point when it comes into scope."),
 # E11 -----------------------------------------------------------------
 "F11.1": ("Unchanged", "High", "Tenancy and patient mapping, backend.", "One paragraph, backend-owned."),
 "F11.2": ("Unchanged", "High", "HL7 provider integration, backend. HISP deferred to BL-43.", "One paragraph, backend-owned."),
 "F11.3": ("Changed", "Medium", "Configuration decides whether doctor engagement is in the app. Open as Q-21.", "Blocked on Q-21."),
 "F11.4": ("Deferred", "High", "User-initiated sharing, all deferred.", "None now."),
 "F11.5": ("Deferred", "High", "Doctor consultation, deferred.", "None now."),
 "F11.6": ("Deferred", "High", "Post-test survey, deferred.", "None now."),
 "F11.7": ("Deferred", "High", "Rating prompt, deferred.", "None now."),
 # E12 -----------------------------------------------------------------
 "F12.1": ("Unchanged", "High", "TLS and certificate pinning as Minuteful.", "One line, plus the new build constraint FR-SEC-014."),
 "F12.2": ("Changed", "High", "Upload works as Minuteful; the application-side result request FR-COM-011 is new.", "The upload states and the new request-and-wait behaviour."),
 "F12.3": ("Unchanged", "High", "Retry and resume on connectivity loss.", "One paragraph: as Minuteful."),
 "F12.4": ("Deferred", "High", "Backend push deferred.", "None now."),
 # E13 -----------------------------------------------------------------
 "F13.1": ("No spec", "High", "Encryption at rest and secure key storage. Platform obligations with no user-facing behaviour.", "None. The requirement is the specification."),
 "F13.2": ("No spec", "High", "Obfuscation, debug disabled, artefact encryption. Build-pipeline work.", "None."),
 "F13.3": ("Unchanged", "High", "Screenshot prevention and app-switcher obscuring. One user-visible consequence, and Q-81 is open against it.", "Short spec: which screens, and what the user sees when blocked."),
 "F13.4": ("No spec", "High", "Error hygiene and input validation. Rules, not behaviour.", "None."),
 "F13.5": ("Unchanged", "High", "Audit events to the backend.", "The event list, jointly with backend."),
 # E14 -----------------------------------------------------------------
 "F14.1": ("Changed", "High", "Mixpanel as Minuteful, but the QACR event taxonomy is new because the steps are new.", "The event list and its properties. Should be written alongside F04.1."),
 "F14.2": ("Unchanged", "High", "No personal or health information, pseudonymous identifier.", "One line: as Minuteful."),
 "F14.3": ("Unchanged", "High", "Analytics failure must not block a test.", "One line: as Minuteful."),
 "F14.4": ("Deferred", "High", "Proxy routing deferred.", "None now."),
 # E15 -----------------------------------------------------------------
 "F15.1": ("No spec", "High", "Parameter control. A process obligation.", "None."),
 "F15.2": ("No spec", "High", "Timing and state verification. Belongs to the test plan.", "None."),
 "F15.3": ("No spec", "High", "Algorithm and release verification. Test plan.", "None."),
 "F15.4": ("No spec", "High", "Off-the-shelf component management. Process.", "None."),
 "F15.5": ("No spec", "High", "Usability testing. Usability Engineering Plan.", "None."),
 "F15.6": ("No spec", "High", "Provider integration testing. Test plan.", "None."),
}

STATUSFILL = {"Unchanged": "1E6B3A", "Changed": "8A6D3B", "New": "9C1F1F",
              "No spec": "6B7789", "Deferred": "5A739B"}

wb = Workbook()


def header(ws, row, hdr, widths, fillin_from=None):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="8A6D3B") if (fillin_from and i >= fillin_from) else HDRF
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BOX
    ws.row_dimensions[row].height = 30


def put(ws, r, vals, fillin_from=None, alt=False, height=None):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=F, size=9, bold=(i == 1))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BOX
        if fillin_from and i >= fillin_from:
            c.fill = FILLIN
        elif alt:
            c.fill = ALT
    if height:
        ws.row_dimensions[r].height = height


# ============================================== READ ME
ws = wb.active
ws.title = "Read me"
ws.sheet_view.showGridLines = False
ws["A1"] = "QACR specs — triage"
ws["A1"].font = Font(name=F, size=16, bold=True, color=INK)
ws["A2"] = "Every feature classified against Minuteful Kidney, so the size of the spec job is known before any spec is written"
ws["A2"].font = Font(name=F, size=10, italic=True, color="666666")

counts = collections.Counter(T[f[0]][0] for f in FEATURES)
rows = [
    ("", ""),
    ("What this is", "My draft classification of all 82 features. It is a starting point for you to correct, not an answer — I have never seen the Minuteful application, so every judgement here is inferred from the requirements, the configuration set, your screenshots and your comments."),
    ("Why it matters", "It converts \"a very big chunk of work\" into a number. On this draft, %d features need a real spec, %d need a paragraph, %d need nothing at all, and %d are deferred." % (counts["New"] + counts["Changed"], counts["Unchanged"], counts["No spec"], counts["Deferred"])),
    ("What to fill in", "Two shaded columns on the Triage tab: correct the status where I have it wrong, and comment where you want to. The Confidence column says where I am guessing — start with the Low and Medium rows, they are where your time is worth most."),
    ("", ""),
    ("The one assumption to check first", "That QACR's timed assay has no Minuteful equivalent. Minuteful tests a reagent strip; QACR has incubation wells, a transfer valve, detection wells and two timed windows. On that basis I marked all six timing features New — which is what makes them the largest piece of genuinely new specification. If Minuteful's develop-and-scan timing is closer than I think, that estimate drops sharply."),
    ("", ""),
    ("Statuses", "Unchanged — recreate as Minuteful; the spec records that, plus any parameter differences. Changed — the mechanism exists but the behaviour differs. New — no Minuteful equivalent; needs a spec written from scratch. No spec — a requirement or a process obligation with nothing to specify. Deferred — out of scope for this version."),
    ("Spec documents", "One per epic, not one per feature. Each has a section per feature, and behaviour statements are numbered by feature — S05.3-2 — so traceability stays at feature granularity. E07 content goes to the copy file instead; E15 process needs no spec. That leaves 13 documents."),
    ("Copy", "No copy text in any spec. Specs cite keys; the copy file holds the strings and is the source of truth. That file is also what FR-TXT-004 already requires: a versioned content set whose version is recorded against each test."),
]
r = 4
for a, b in rows:
    ws.cell(row=r, column=1, value=a).font = Font(name=F, size=10, bold=True, color=INK)
    c = ws.cell(row=r, column=2, value=b)
    c.font = Font(name=F, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if a == "The one assumption to check first":
        ws.cell(row=r, column=1).fill = FILLIN
        c.fill = FILLIN
    ws.row_dimensions[r].height = 14 if not b else max(28, 13 * (len(b) // 95 + 1))
    r += 1
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 108

r += 1
ws.cell(row=r, column=1, value="Counts on this draft").font = Font(name=F, size=10, bold=True, color=INK)
r += 1
for k in ["New", "Changed", "Unchanged", "No spec", "Deferred"]:
    ws.cell(row=r, column=1, value=k).font = Font(name=F, size=10)
    c = ws.cell(row=r, column=2, value=counts[k])
    c.font = Font(name=F, size=10, bold=True, color=STATUSFILL[k])
    r += 1
ws.cell(row=r, column=1, value="Total").font = Font(name=F, size=10, bold=True)
ws.cell(row=r, column=2, value="=SUM(B%d:B%d)" % (r - 5, r - 1)).font = Font(name=F, size=10, bold=True)

# ============================================== TRIAGE
ws = wb.create_sheet("Triage")
ws.sheet_view.showGridLines = False
ws["A1"] = "Correct the status where I have it wrong. Confidence says where I am guessing — Low and Medium first."
ws["A1"].font = Font(name=F, size=10, italic=True, color="666666")
ws.merge_cells("A1:K1")
header(ws, 2,
       ["Ref.", "Epic", "Feature", "M", "Domain", "Status", "Conf.", "Why I think so", "What the spec needs",
        "YOUR STATUS", "YOUR COMMENT"],
       [7, 6, 26, 8, 17, 11, 7, 46, 44, 12, 30], fillin_from=10)

dv = DataValidation(type="list", formula1='"Unchanged,Changed,New,No spec,Deferred"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)

r = 3
for fid, epic, etitle, fname, ms, dm, nfr, nbl in FEATURES:
    st, conf, why, needs = T[fid]
    put(ws, r, [fid, epic, fname, ms, dm, st, conf, why, needs, "", ""], fillin_from=10, height=48)
    ws.cell(row=r, column=6).font = Font(name=F, size=9, bold=True, color=STATUSFILL[st])
    ws.cell(row=r, column=7).font = Font(name=F, size=9, bold=(conf != "High"),
                                         color="9C1F1F" if conf == "Low" else ("8A6D3B" if conf == "Medium" else "666666"))
    dv.add(ws.cell(row=r, column=10))
    r += 1
ws.freeze_panes = "A3"
ws.auto_filter.ref = "A2:K%d" % (r - 1)

# ============================================== SPEC PLAN
ws = wb.create_sheet("Spec plan")
ws.sheet_view.showGridLines = False
ws["A1"] = "The 13 spec documents, in the order I would write them. Milestone 1 first, because mid-October governs."
ws["A1"].font = Font(name=F, size=10, italic=True, color="666666")
ws.merge_cells("A1:G1")
header(ws, 2, ["Order", "Spec", "Covers", "Features", "Weight", "Why this order", "YOUR COMMENT"],
       [7, 9, 34, 30, 12, 44, 30], fillin_from=7)

PLAN = [
 ("1", "S-E05", "Assay Timing Control", "F05.1 to F05.6 — six, all new", "Heaviest",
  "Entirely new, milestone 1, and safety-critical: the application is the only thing enforcing either window. Also the one where my assumptions are weakest, so it needs your input earliest."),
 ("2", "S-E06", "Scan Capture and Image Validation", "F06.1 to F06.6 — six, all changed", "Heavy",
  "Milestone 1, largest requirement count, and the two-stage pipeline needs writing down once, properly. Depends on the QACR IVTS document for thresholds."),
 ("3", "S-E04", "Guided Test Flow Engine", "F04.1 to F04.8 — three new, three changed, two unchanged", "Heavy",
  "Milestone 1. The chat sequencer is the most repeated interaction in the product, and scan practice and the waiting card hang off it."),
 ("4", "S-E09", "Result Presentation", "F09.1, F09.2, F09.4", "Medium",
  "Milestone 1. Small but consequential, and it settles what demonstration mode shows."),
 ("5", "S-E12", "Connectivity and Data Transmission", "F12.1 to F12.3", "Light",
  "Milestone 1 for upload and the new result request; the rest is as Minuteful."),
 ("6", "S-E03", "Kit Identification", "F03.1, F03.2", "Light",
  "Milestone 3, but small and self-contained."),
 ("7", "S-E02", "Identity, Authentication and Consent", "F02.1 to F02.8", "Medium",
  "Milestone 3. Mostly recreation; the date-of-birth change and onboarding need writing."),
 ("8", "S-E01", "App Foundation and Device Readiness", "F01.1 to F01.9, less F01.7", "Light",
  "Milestone 3, almost all unchanged. The alert pattern is the only part worth care."),
 ("9", "S-E10", "Results Center, Menu and Support", "F10.1 to F10.5", "Light",
  "Milestone 3 to 4. Recreation, except the results-center split."),
 ("10", "S-E14", "Analytics", "F14.1 to F14.3", "Light",
  "Milestone 3. The event taxonomy should be written alongside S-E04, not after it."),
 ("11", "S-E13", "Security and Privacy", "F13.3, F13.5 only", "Light",
  "Milestone 3. Most of the epic needs no spec; only screenshot protection and the audit event list do."),
 ("12", "S-E08", "Result Computation", "F08.2, F08.4", "Light",
  "Milestone 2 to 3, but F08.1 and F08.3 belong to the algorithm team's own document, referenced not restated."),
 ("13", "S-E11", "Result Routing", "F11.1, F11.2, F11.3", "Light",
  "Milestone 3, backend-owned, and F11.3 is blocked on Q-21."),
]
r = 3
for row in PLAN:
    put(ws, r, list(row) + [""], fillin_from=7, height=46, alt=(r % 2 == 0))
    ws.cell(row=r, column=5).font = Font(name=F, size=9, bold=True,
        color={"Heaviest": "9C1F1F", "Heavy": "8A6D3B", "Medium": "1F3864", "Light": "1E6B3A"}[row[4]])
    r += 1
ws.freeze_panes = "A3"

# ============================================== COPY FILE
ws = wb.create_sheet("Copy file")
ws.sheet_view.showGridLines = False
ws["A1"] = "How I would set the copy file up, following your decision that copy does not live in specs. Correct or replace with the Minuteful scheme."
ws["A1"].font = Font(name=F, size=10, italic=True, color="666666")
ws.merge_cells("A1:D1")
header(ws, 2, ["", "Proposal", "Why", "YOUR COMMENT"], [4, 46, 50, 32], fillin_from=4)

COPY = [
 ("One", "Two files, or one file with two top-level groups: chat items and static items, as you described.",
  "Matches how the app consumes it and how you already think about it. Chat items are the flow bubbles; static items are screens, alerts and labels."),
 ("Two", "Every string has a stable key. Specs cite the key, never the text.",
  "This is what stops copy being duplicated and going stale. It also means copy can be rewritten without touching a spec, and reviewed as one body — which is what FR-TXT-004 asks for."),
 ("Three", "Keys carry the feature they belong to, so the copy file is traceable the same way everything else is.",
  "A build check then works in both directions: every key a spec cites exists, and every key is cited by something or explicitly marked unused."),
 ("Four", "The copy file is the versioned content set FR-TXT-004 requires, and its version is recorded against each test record.",
  "Turns an existing requirement into a real artefact rather than an aspiration. Also gives the Risk Analysis something concrete to point at for instruction content."),
 ("Five", "Flesch-Kincaid computed per string at build time, failing the build above grade 6.",
  "FR-TXT-001 sets a grade-6 ceiling on all user-facing text. Reviewing that by eye across several hundred strings is not realistic; computing it is a few lines and turns a manual regulatory check into a build failure. Note the open question Q-22 on whether mandated regulatory wording is exempt."),
 ("Six", "Component names checked against the approved glossary set.",
  "FR-TXT-002 bars internal engineering names from user-facing text. A word list check catches Transfer Valve or Peel-Off Seal appearing in a string where the approved name should be."),
 ("Seven", "Voice script fields alongside the display text, where they differ.",
  "FR-FLW-008 reads the flow aloud and FR-IMG-024 speaks scan guidance. Spoken wording is often not the written wording, so the file needs to hold both rather than forcing one to serve."),
]
r = 3
for row in COPY:
    put(ws, r, list(row) + [""], fillin_from=4, height=52, alt=(r % 2 == 0))
    r += 1
ws.freeze_panes = "A3"

# ============================================== WHAT I NEED
ws = wb.create_sheet("What I need")
ws.sheet_view.showGridLines = False
ws["A1"] = "Inputs, in the order they unblock work."
ws["A1"].font = Font(name=F, size=10, italic=True, color="666666")
ws.merge_cells("A1:E1")
header(ws, 2, ["", "Input", "Unblocks", "Notes", "YOUR COMMENT"], [4, 34, 26, 46, 30], fillin_from=5)

NEED = [
 ("One", "This triage, corrected.", "Everything",
  "Especially the six timing rows and the four I marked Low or Medium confidence. If Minuteful's timing model is closer to QACR's than I assume, the plan changes shape."),
 ("Two", "The Minuteful copy file, or its structure.", "The copy scheme and every spec",
  "I would rather adopt your key naming than invent one and have to migrate."),
 ("Three", "Screenshots for the features marked Unchanged.", "The light specs",
  "A screenshot per screen is usually enough for an Unchanged feature: it tells me the states, the controls and the copy keys without a written spec."),
 ("Four", "The QACR IVTS specification, and the algorithm document if one exists.", "S-E06, and the F08 rows",
  "Thresholds, the frame recipe and the boundary conditions should be referenced from those documents, not restated in a spec where they would go stale."),
 ("Five", "Any written specs you do have.", "Whichever features they cover",
  "The lobby 4 spec is already in the source documents folder. Others would save drafting and, more importantly, tell me your house style."),
 ("Six", "A decision on whether I get the repositories.", "Verification, not drafting",
  "Not needed to start. Useful at the end of each spec as a check that what I wrote matches what the code does — with the caveat that where they disagree it is a question for you, not automatically a correction to the spec."),
]
r = 3
for row in NEED:
    put(ws, r, list(row) + [""], fillin_from=5, height=52, alt=(r % 2 == 0))
    r += 1
ws.freeze_panes = "A3"

for w in wb:
    w.page_setup.orientation = "landscape"
    w.page_setup.fitToWidth = 1
    w.page_setup.fitToHeight = 0
    w.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUT)
print("wrote", OUT)
print("features:", len(FEATURES), " statuses:", dict(counts))
