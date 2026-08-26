#!/usr/bin/env python3
"""Builds the conflict review worksheet from the engineering analysis, with every
claim already checked against QACR-APP-FR-01 Rev 1.6."""
import json, subprocess, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = sys.argv[1]

INK = "1F3864"
HDRF = PatternFill("solid", fgColor=INK)
ALT = PatternFill("solid", fgColor="F2F5FA")
FILLIN = PatternFill("solid", fgColor="FFF6D6")
GRPF = PatternFill("solid", fgColor="DCE3EF")
THIN = Side(style="thin", color="BFC9DA")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F = "Arial"


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=F, size=9, bold=True, color="FFFFFF")
        cell.fill = HDRF
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BOX
    ws.row_dimensions[row].height = 30


def put(ws, r, vals, widths=None, fillin_from=None, alt=False, bold_first=True):
    for i, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = Font(name=F, size=9, bold=(bold_first and i == 1), color="000000")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BOX
        if fillin_from and i >= fillin_from:
            cell.fill = FILLIN
        elif alt:
            cell.fill = ALT


wb = Workbook()

# ===========================================================  READ ME
ws = wb.active
ws.title = "Read me"
ws.sheet_view.showGridLines = False
ws["A1"] = "QACR functional requirements — conflict review"
ws["A1"].font = Font(name=F, size=16, bold=True, color=INK)
ws["A2"] = "Worksheet for resolving the 19 dependency and timeline issues raised against QACR-APP-FR-01 Rev 1.5"
ws["A2"].font = Font(name=F, size=10, italic=True, color="666666")

rows = [
    ("", ""),
    ("What this is", "One row per conflict on the Conflicts tab. Read column F for the conflict in a line, then fill the two shaded columns at the right — J, Your decision, and K, Your comment. Nothing has been changed in the requirements document."),
    ("What to fill in", "Only the shaded columns. Your decision has a dropdown; Your comment is free text. A comment such as \"mock results are fine for the demo, add a requirement\" is enough for me to draft the change."),
    ("Three ways a conflict resolves", "Change a priority · Add a requirement that scopes the earlier milestone down (for example: the demonstration build presents a fixed result payload rather than a computed one) · Withdraw the requirement. Two issues resolve instead by naming an owner and a date."),
    ("Rows marked \"mock-scope candidate\"", "Column H flags the five conflicts where adding a narrower requirement for the earlier milestone is a live option, rather than moving a priority. Those are the ones your mock-results idea applies to: CR-02, CR-04, CR-07, CR-08 and CR-21."),
    ("Other tabs", "Doc corrections — nine stale references and text defects that need no product decision, only your go-ahead. Analysis checked — the six places where I verified, qualified or disagreed with the analysis, so you know what to trust."),
    ("", ""),
    ("How much of the analysis I verified", "All 82 priorities it quotes match Rev 1.6 exactly. All five missing identifiers are genuinely absent. Every stale note it quotes is stale. The transcription is reliable; the reasoning needed qualifying in six places, on the Analysis checked tab."),
    ("Where it agrees with our own work", "Seven of these conflicts are already in section 4 of QACR-APP-EPIC-01, found independently by grouping requirements into features. Two conflicts here are ones we found and it did not; twelve are ones it found and we did not."),
    ("", ""),
    ("ONE THING TO CONFIRM FIRST", "The analysis assumes M1 = mid-October 2026, M2 = early-December, M3 and M4 = early-January 2027, and that M3 and M4 share a date. None of that is in our requirements document. It drives the framing of half these issues — the rule \"a must-have cannot depend on a nice-to-have\" holds regardless of dates, but please confirm or correct the calendar."),
]
r = 4
for a, b in rows:
    ws.cell(row=r, column=1, value=a).font = Font(name=F, size=10, bold=True, color=INK)
    c = ws.cell(row=r, column=2, value=b)
    c.font = Font(name=F, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if a.isupper() and a:
        ws.cell(row=r, column=1).fill = FILLIN
        c.fill = FILLIN
    ws.row_dimensions[r].height = 14 if not b else max(28, 13 * (len(b) // 95 + 1))
    r += 1

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 108

# live counters
r += 1
ws.cell(row=r, column=1, value="Progress").font = Font(name=F, size=10, bold=True, color=INK)
r += 1
for label, formula in [
    ("Conflicts awaiting a decision", '=COUNTBLANK(Conflicts!J3:J23)'),
    ("Conflicts decided", '=21-COUNTBLANK(Conflicts!J3:J23)'),
    ("Corrections awaiting go-ahead", '=COUNTBLANK(\'Doc corrections\'!G3:G11)'),
]:
    ws.cell(row=r, column=1, value=label).font = Font(name=F, size=10)
    c = ws.cell(row=r, column=2, value=formula)
    c.font = Font(name=F, size=10, bold=True, color=INK)
    r += 1

# ===========================================================  CONFLICTS
ws = wb.create_sheet("Conflicts")
ws.sheet_view.showGridLines = False
hdr = ["Ref", "Their ref", "Group", "Committed / earlier side", "Later, weaker or missing side",
       "The conflict, in one line", "Feature", "Resolution options open to you",
       "My recommendation", "YOUR DECISION", "YOUR COMMENT"]
widths = [6, 9, 11, 25, 25, 62, 8, 42, 42, 20, 34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws["A1"] = "19 issues from the engineering analysis, plus 2 found in our own consistency pass. Every priority quoted has been checked against Rev 1.6."
ws["A1"].font = Font(name=F, size=10, italic=True, color="666666")
ws.merge_cells("A1:K1")
style_header(ws, 2, len(hdr))
for i, h in enumerate(hdr, start=1):
    ws.cell(row=2, column=i, value=h)
style_header(ws, 2, len(hdr))
for col in ("J", "K"):
    ws.cell(row=2, column=10 if col == "J" else 11).fill = PatternFill("solid", fgColor="8A6D3B")

CONFLICTS = [
    # ref, their ref, group, early side, late side, one-line, feature, options, recommendation
    ("CR-01", "DT15-07", "Priority inversion",
     "10+ reqs at M1 and M3 realised by config keys; FR-PLT-005 (3), FR-KIT-007 (3) say \"configurable\" in their own text",
     "FR-CFG-001 (4), FR-CFG-002 (4)",
     "Configuration retrieval is nice-to-have, yet requirements at M1 and M3 depend on configuration being there, and two must-haves require a configurable value in their own wording. If M4 slips there is no requirement that configuration is retrieved at all.",
     "F01.9",
     "Move FR-CFG-001 and FR-CFG-002 to 3; or state for each dependent requirement that the value is fixed in the build for this version",
     "Move both to 3. The retrieval mechanism has to exist before anything can read it, and 10+ requirements read it."),
    ("CR-02", "DT15-13", "Priority inversion",
     "FR-TIM-005, 006, 007, 008, 009, 011 (all 1)",
     "FR-TIM-001 (3), FR-TIM-002 (3), FR-TIM-003 (3), FR-TIM-012 (3); BL-24 deferred",
     "Six M1 requirements enforce the two safety-critical timing windows, but the requirement that defines start event, minimum duration, completion window and expiry is M3, as is the rule for recomputing elapsed time after a screen lock and the message shown when a test is invalidated on time.",
     "F05.1–F05.6",
     "Move FR-TIM-001, FR-TIM-003 and FR-TIM-012 to 1; or add a demonstration-scope requirement with provisional timing values  [mock-scope candidate]",
     "Move all three to 1. Without FR-TIM-003 the demo's timers restart on backgrounding, so the demo is not evidence that either window is enforced."),
    ("CR-03", "DT15-06", "Priority inversion",
     "FR-PRT-001 (3)",
     "FR-ACC-001, 002, 004, 005, 006, 009 (all 4); BL-25, BL-26, BL-27 deferred",
     "The results portal is committed for M3 and every control protecting access to it is M4 — the PIN gate, the prompt, the numeric restriction, the lockout, the setup window and the inactivity timeout. All six trace to the threat analysis, SPTA 1.4, 1.7 and 2.1.",
     "F10.1–F10.3",
     "Move section 18 to 3; or hold FR-PRT-001 back to 4 so screen and protection ship together; or accept and have regulatory record it",
     "Move section 18 to 3. Note the exposure is narrower than the analysis states: the backend enforces the PIN, the lockout and the 24-hour window server-side."),
    ("CR-04", "DT15-08", "Priority inversion",
     "FR-IMG-004, 005, 006, 007, 008, 011, 014, FR-COM-008, FR-RDY-008, FR-RDY-009, FR-STA-007, FR-TIM-006 (all 1)",
     "FR-TXT-020 (3)",
     "Twelve M1 requirements block the user and show a message; the standard that governs what every blocking message must say is M3. Either the October copy is written twice, or the October build ships messages that do not meet the standard.",
     "F07.4",
     "Move FR-TXT-020 to 1; or state that October messages are provisional and rewritten at M3  [mock-scope candidate]",
     "Move FR-TXT-020 to 1. It is one requirement and it is the cheapest fix on this list; writing 25 messages twice is not."),
    ("CR-05", "DT15-09", "Priority inversion",
     "FR-SUP-003 (3), FR-TXT-020 (3), FR-RDY-011 (3), FR-AUT-016 (3), FR-PRT-004 (3), FR-KIT-001 (3)",
     "FR-SUP-002 (4); BL-40 deferred",
     "Six must-haves direct the user to contact support, and the requirement that actually provides a telephone number and an email address is nice-to-have. SPTA 5.2 relies on support offering workarounds during an outage.",
     "F10.5",
     "Move FR-SUP-002 to 3; or name which single support route the must-have messages may cite in this version",
     "Move FR-SUP-002 to 3. A threat-analysis control should not rest on a nice-to-have, and six requirements point at it."),
    ("CR-06", "DT15-10", "Priority inversion",
     "FR-KIT-002 (3), FR-KIT-007 (3)",
     "FR-KIT-004 (4); BL-23 deferred",
     "Detection that a kit was already used is M3 and the new-kit alert is M3, but the requirement stating what happens when a used kit is presented — the only thing that stops a reused kit producing a result — is M4.",
     "F03.1, F03.2",
     "Move FR-KIT-004 to 3; or state that in this version reuse is detected by the backend check alone and say what the app does with that answer",
     "Move FR-KIT-004 to 3. FR-KIT-007 describes itself as the complement to FR-KIT-004, so committing the complement and not the control is backwards."),
    ("CR-07", "DT15-11", "Priority inversion",
     "FR-RES-001, 002, 003, 005 (all 1)",
     "FR-ALG-010 (2), FR-ALG-012 (2), FR-TXT-021 (3); and FR-RES-004 (2) against FR-ALG-011 (3)",
     "The result screens are M1 while the payload contract that distinguishes a valid result from an invalid one is M2, and the invalid-test messaging is M3. The invalid-reason enumeration everything rests on is not defined anywhere.",
     "F09.1, F09.2, F08.4",
     "Add a demonstration-scope requirement: at M1 the application presents a fixed or mock result payload and produces no clinical result; or align FR-RES-005, FR-ALG-010, FR-ALG-012 and FR-TXT-021 on one priority  [mock-scope candidate]",
     "This is your mock-results case. Add the demo-scope requirement and leave the priorities alone — it is the honest description of what M1 actually is."),
    ("CR-08", "DT15-12", "Priority inversion",
     "FR-ALG-001 (1)",
     "FR-ALG-002 (3), FR-ALG-003 (no priority)",
     "FR-ALG-001 at M1 requires an algorithm version \"approved for the applicable system configuration\", but the requirement maintaining that approval mapping is M3 and the one enforcing an approved combination has no priority. In October nothing defines what approved means.",
     "F08.1, F08.2",
     "Move FR-ALG-002 to 1; or mark the \"approved\" clause as not verifiable at M1 via the same demonstration-scope requirement as CR-07  [mock-scope candidate]",
     "Cover it with the CR-07 demo requirement rather than moving FR-ALG-002. If the demo shows a mock result there is no algorithm version to approve."),
    ("CR-09", "DT15-14", "Priority inversion",
     "Everything at 1 and 3; FR-SHR-001 to 008 (all 3)",
     "FR-LCM-006 (4), FR-LCM-008 (4), FR-LCM-016 (4)",
     "Three verification requirements are nice-to-have while what they verify is must-have. FR-LCM-006 requires every parameter to have a value and rationale before design verification, so a precondition of verification sits behind the things it is a precondition for.",
     "F15.1, F15.2, F15.6",
     "Move FR-LCM-006 to 3; decide separately whether provider integration testing can be nice-to-have while provider routing is must-have",
     "Move FR-LCM-006 to 3. It is the requirement that closes every open parameter value in the register, so it cannot be optional."),
    ("CR-10", "DT15-15", "Priority inversion",
     "FR-AUT-001 (3), FR-PRT-002 (3)",
     "FR-CNS-001, 002, 003 (all 4), FR-PRT-005 (4); BL-03 deferred",
     "Phone-number entry is committed for M3 and every consent obligation attached to it is M4, so the committed login could ship with no acknowledgement checkbox and no route to the policy documents.",
     "F02.7, F10.4",
     "Move FR-CNS-001 to 003 and FR-PRT-005 to 3; or accept and have regulatory record having accepted it",
     "Move all four to 3. Collecting a phone number without recording acceptance of the privacy policy is not a scheduling question."),
    ("CR-11", "DT15-16", "Priority inversion",
     "FR-PLT-005 (3)",
     "FR-PLT-004 (4)",
     "FR-PLT-005 narrows its own scope by handing model-level exclusion to FR-PLT-004, which is nice-to-have, so excluding a specific unsupported phone model is committed by nothing.",
     "F01.1, F01.2",
     "Move FR-PLT-004 to 3; or state that model-level exclusion is out of scope and the risk is carried at run time by FR-RDY-002 and FR-RDY-003",
     "Move FR-PLT-004 to 3. It is a store-console setting rather than development work, so the cost of committing it is near zero."),
    ("CR-12", "DT15-04", "No priority",
     "FR-IMG-004 to 011 (all 1) depend on FR-CAM-002; FR-ALG-001 (1) on FR-ALG-003",
     "FR-CAM-002, FR-ALG-003, FR-ALG-008, FR-AUT-010, FR-AUT-012, FR-SHR-011 — all TBD",
     "Six requirements still carry no priority, so they belong to no milestone and no owner. FR-CAM-002 is the urgent one: it applies the camera adjustments raised by eight priority-1 boundary conditions and is the only requirement in its section without a priority.",
     "F06.1, F08.2, F08.3, F02.4, F11.3",
     "A priority for each, or an explicit withdrawal. If only one is settled this round, settle FR-CAM-002",
     "Settle FR-CAM-002 at 1 now. The other five are already tracked as R.1 in the register; this row adds what each one blocks."),
    ("CR-13", "DT15-05", "No priority",
     "—",
     "FR-AUT-010 (TBD), FR-AUT-012 (TBD)",
     "Both notes say \"not wanted for the clinical study but required for the submission\", which is a schedule position the four-milestone scale cannot express. The feature-flag suggestion answers how, not by when.",
     "F02.4",
     "Confirm the submission date is M3, in which case both are priority 3 behind a flag; or add a fifth milestone with its own date",
     "Confirm whether the submission has a date distinct from the clinical study. If it does not, both are 3 behind a flag and the note should say so."),
    ("CR-14", "DT15-01", "Cannot be satisfied",
     "—",
     "FR-AUT-009 (4); BL-20 and BL-21 deferred",
     "With shared phone numbers deferred, a phone number maps to exactly one user, so FR-AUT-009's condition is always true. It is satisfied by writing no code and cannot be verified as doing anything.",
     "F02.4",
     "Withdraw FR-AUT-009 to Appendix D; or state that date of birth is part of login in this version and restate it",
     "Decide first whether date of birth is part of login at all in this version — CR-14, CR-15 and CR-16 all turn on that one answer."),
    ("CR-15", "DT15-02", "Cannot be satisfied",
     "FR-ACC-009 (4)",
     "FR-AUT-008 (4), FR-AUT-012 (TBD), BL-20 deferred; FR-AUT-009 forbids it at login",
     "FR-ACC-009 requires the date of birth to be submitted \"again\" after a portal timeout, but nothing scheduled captures it the first time: FR-AUT-009 forbids asking at login, FR-AUT-012 has no priority, BL-20 is deferred.",
     "F10.2, F02.4",
     "Name the requirement that captures date of birth and give it a priority; or restate FR-ACC-009 to re-enter the portal by PIN instead",
     "PIN re-entry is the simpler answer and the PIN already exists in section 18. It also removes the dependency on the two TBD requirements."),
    ("CR-16", "DT15-03", "Cannot be satisfied",
     "RA 6.4 — results attributed to the wrong patient",
     "Appendix D names FR-AUT-007, 009, 010, 012, 013; BL-20 names FR-AUT-009, 013, 014 and FR-SHR-003",
     "Two appendices of the same document give different lists of what traces to RA 6.4, and most of what Appendix D names is deferred or unscheduled. Only FR-AUT-013 has both a milestone and a reachable trigger.",
     "F02.4, F02.5",
     "One list, containing only requirements that are in scope and have a priority",
     "Adopt the BL-20 list — FR-AUT-009 excepted, per CR-14 — and correct the Appendix D row. Follows automatically once CR-14 is decided."),
    ("CR-17", "DT15-17", "Owner or date",
     "FR-RES-001 (1), FR-RES-005 (1), FR-PRT-001 (3), FR-ALG-012 (2), FR-TXT-021 (3); all of section 12",
     "No owner for the results endpoint, the payload schema or the invalid-reason enumeration; no date for the timing-flex study",
     "Two inputs the schedule rests on have no owner and no date. Separately, no requirement in the document covers the application-side act of fetching a result — FR-RES-001 presumes receipt.",
     "F08.4, F05.1",
     "Name an owner for the endpoint contract and the enumeration; set a date for the timing-flex study; decide whether to add an application-side fetch requirement",
     "The timing-flex study date is the long-lead item — ask for it first. The missing fetch requirement I would add regardless; it is a real gap, not a wording point."),
    ("CR-18", "DT15-18", "Owner or date",
     "FR-LCM-017 (3), and the Appendix D exclusion of FR-LCM-001",
     "Software Development Plan — no owner, no date, not in the register",
     "FR-LCM-017 scopes release testing by reference to a Software Development Plan, and the exclusion of FR-LCM-001 puts the Class B development obligation there too. Nothing in the document set says who owns it or when it lands.",
     "F15.3",
     "Name an owner and a date; or bring the release-testing scope into this document",
     "Name an owner. Two load-bearing obligations point at a document nobody is recorded as writing."),
    ("CR-19", "DT15-19", "Bookkeeping",
     "—",
     "FR-IMG-018, FR-IMG-019, FR-SHR-009, FR-SHR-010, FR-TIM-013 — unaccounted; Appendix D structure; BL numbering",
     "Five identifiers appear in none of the three places an identifier may live, Appendix D mixes exclusions with consolidations and superseded wordings, and BL numbers have moved between revisions while two closed decisions cite them.",
     "—",
     "See the Doc corrections tab — nine items, each needing only your go-ahead",
     "I verified all five identifiers are genuinely absent. Handle it all on the Doc corrections tab; no priority decision here."),
    ("CR-20", "DT15-20", "Confirm only",
     "RA 4.17",
     "FR-RDY-004 (3), sole remaining control after FR-PLT-007 was withdrawn",
     "Withdrawing automatic update leaves RA 4.17 resting on the version check alone, and that check is M3, so until January an unsupported application version is blocked only by the backend's own validation.",
     "F01.2",
     "Confirm RA 4.17 is intended to rest on the version check alone, and that the Risk Analysis drops any reference to automatic update at its next revision",
     "Confirm and move on. FR-RDY-004's note already records enforcement on both sides, so this is close to closed."),
    ("CR-21", "ours", "Priority inversion",
     "FR-COM-003 (1), FR-COM-004 (1), FR-COM-005 (1)",
     "FR-COM-001 (3), FR-COM-002 (3)",
     "Test data is uploaded at M1 while TLS enforcement and certificate pinning are M3, so a demonstration build transmits over whatever the platform negotiates. Not in the engineering analysis; found in our own consistency pass.",
     "F12.1, F12.2",
     "Accept, with a requirement that no patient data passes through a demonstration build  [mock-scope candidate]; or move FR-COM-001 and FR-COM-002 to 1",
     "Accept with the demo-scope requirement. Demonstration builds have a way of being reused, so the constraint is worth writing down."),
]

DECISIONS = ("Change priority,Add requirement,Withdraw,Assign owner and date,"
             "Accept as-is,Restate wording,Need to discuss")
dv = DataValidation(type="list", formula1='"%s"' % DECISIONS, allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)

r = 3
last_group = None
for c in CONFLICTS:
    put(ws, r, list(c) + ["", ""], fillin_from=10, alt=(c[2] == last_group and r % 2 == 0))
    ws.cell(row=r, column=3).fill = GRPF
    ws.cell(row=r, column=3).font = Font(name=F, size=8, bold=True, color=INK)
    ws.cell(row=r, column=1).font = Font(name=F, size=9, bold=True, color=INK)
    ws.cell(row=r, column=6).font = Font(name=F, size=9)
    for col in (10, 11):
        ws.cell(row=r, column=col).fill = FILLIN
    dv.add(ws.cell(row=r, column=10))
    ws.row_dimensions[r].height = 74
    last_group = c[2]
    r += 1

# example row so the expected format is unambiguous
ws.cell(row=r + 1, column=1, value="Example of what a filled row looks like:").font = Font(name=F, size=9, italic=True, color="8A6D3B")
put(ws, r + 2, ["CR-07", "DT15-11", "Priority inversion", "—", "—",
                "(the conflict text stays as it is — you only fill the last two columns)", "—", "—", "—",
                "Add requirement", "Mock results are fine for the demo. Add a requirement that M1 shows a fixed payload, leave the priorities alone."],
    fillin_from=10)
for col in (10, 11):
    ws.cell(row=r + 2, column=col).fill = PatternFill("solid", fgColor="FFEFC0")
    ws.cell(row=r + 2, column=col).font = Font(name=F, size=9, italic=True, color="8A6D3B")
ws.row_dimensions[r + 2].height = 46

ws.freeze_panes = "A3"
ws.auto_filter.ref = "A2:K%d" % (len(CONFLICTS) + 2)

# ===========================================================  DOC CORRECTIONS
ws = wb.create_sheet("Doc corrections")
ws.sheet_view.showGridLines = False
ws["A1"] = "Stale references and text defects. No product decision needed — each needs only your go-ahead, then I fix it in Rev 1.7."
ws["A1"].font = Font(name=F, size=10, italic=True, color="666666")
ws.merge_cells("A1:H1")
hdr = ["Ref", "Where", "What it says now", "What is wrong", "What I would change it to", "Source", "GO AHEAD?", "YOUR COMMENT"]
widths = [7, 22, 46, 44, 46, 10, 12, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(hdr, start=1):
    ws.cell(row=2, column=i, value=h)
style_header(ws, 2, len(hdr))
for col in (7, 8):
    ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor="8A6D3B")

CORR = [
    ("DC-01", "FR-AUT-009, Notes",
     "\";The converse of FR-AUT-007\"",
     "Stray leading semicolon, and it points at an identifier that is no longer in scope — FR-AUT-007 is now BL-20.",
     "\"The converse of BL-20\" — or the whole requirement goes, if CR-14 is decided as a withdrawal.",
     "DT15-01"),
    ("DC-02", "FR-PRT-004, Notes",
     "\"Adds menu access to the support routes required by FR-SUP-002 and FR-SUP-004\"",
     "FR-SUP-004 was withdrawn and is now BL-40, so a must-have note points at a retired identifier.",
     "\"Adds menu access to the support route required by FR-SUP-002. In-application support chat is deferred as BL-40.\"",
     "DT15-09"),
    ("DC-03", "Appendix D, FR-AUT-009 row",
     "\"Rewritten; the identification case is stated in FR-AUT-007 and FR-AUT-009 states the converse\"",
     "Justifies a surviving requirement by reference to one that has since been deferred. The rewrite record is stale.",
     "Name BL-20 rather than FR-AUT-007, and note that the identification case is deferred.",
     "DT15-01"),
    ("DC-04", "Appendix D — FR-LCM-014, FR-LCM-015 and the FR-ALG-008 superseded row",
     "\"FR-TXT-001 remains at priority 0\", \"FR-TXT-002 remains at priority 0\", \"each retain at least one other requirement at priority 0\"",
     "Priority 0 does not exist in the milestone scale — it is left over from the earlier 0-to-3 scheme. FR-TXT-001 and FR-TXT-002 are both priority 2. The FR-ALG-008 row uses the phrase to argue that certain Risk Analysis rows are covered, so it is not merely cosmetic.",
     "\"remains at priority 2\" in the first two; in the FR-ALG-008 row, state the actual milestone of each retained requirement.",
     "DT15-19"),
    ("DC-05", "Appendix E, BL-13",
     "\"FR-ACC-005 states the behaviour; the limit itself is deferred\"",
     "The limit was decided at Rev 1.4 — Q-07, five consecutive attempts — and landed in FR-ACC-004 and FR-ACC-005. The backlog still reads as though it is outstanding work.",
     "Close BL-13 against FR-ACC-005 and remove it from Appendix E, recording the closure in the decision log.",
     "DT15-06"),
    ("DC-06", "Appendix D, FR-CFG-005 row",
     "\"the readiness thresholds in section 5 are stated as fixed in the application rather than configurable\"",
     "True of FR-RDY-007 only. FR-RDY-002 names staticData.supportedHardwareConfig, FR-RDY-008 names lowDiskSpaceSize, FR-RDY-004 names iosMinAppVersion and forceUpdateOnPreLogin — all configuration. The claim is broader than the requirements support, and it is used to justify not including FR-CFG-005.",
     "Narrow the claim to FR-RDY-007, and state which readiness values are configuration-supplied. Revisit whether FR-CFG-005 should be excluded once CR-01 is decided.",
     "DT15-07"),
    ("DC-07", "Appendix D, structure",
     "19 rows presented as one list",
     "Three different kinds of row with no separation: 11 genuine exclusions, 6 consolidations into a requirement still in scope, and 2 superseded wordings. Only the first group is an exclusion.",
     "Split into three labelled groups under the existing appendix. The two superseded wordings are already suffixed \"(earlier wording)\", so they are distinguishable, but the heading should say so.",
     "DT15-19"),
    ("DC-08", "Identifier record",
     "FR-IMG-018, FR-IMG-019, FR-SHR-009, FR-SHR-010, FR-TIM-013 appear nowhere",
     "I checked all three places an identifier may live — the 184 in scope, the 19 Appendix D rows, and every \"formerly FR-\" note in the backlog. These five are in none of them, and every other number in every prefix is accounted for.",
     "Either add a line to Appendix D recording each as withdrawn, or add a note that the numbering skips them. I cannot tell which without knowing whether they ever carried text.",
     "DT15-19"),
    ("DC-09", "Appendix E, numbering",
     "40 entries running BL-01 to BL-41 with BL-14 absent; FR-SUP-004 was BL-46 in an earlier analysis and is BL-40 now",
     "BL references have moved between revisions, and two closed decisions — Q-11 and Q-12 — record their outcomes against BL-03 and BL-23. A decision whose record points at a moving reference is hard to audit.",
     "State in section 2 that BL references are stable from Rev 1.6 onward and are not reissued, matching the rule already stated for requirement identifiers.",
     "DT15-19"),
]
dv2 = DataValidation(type="list", formula1='"Yes,No,Discuss"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv2)
r = 3
for c in CORR:
    put(ws, r, list(c) + ["", ""], fillin_from=7)
    ws.cell(row=r, column=1).font = Font(name=F, size=9, bold=True, color=INK)
    for col in (7, 8):
        ws.cell(row=r, column=col).fill = FILLIN
    dv2.add(ws.cell(row=r, column=7))
    ws.row_dimensions[r].height = 62
    r += 1
ws.freeze_panes = "A3"

# ===========================================================  ANALYSIS CHECKED
ws = wb.create_sheet("Analysis checked")
ws.sheet_view.showGridLines = False
ws["A1"] = "What I verified against Rev 1.6, and the six places the analysis needs qualifying. Nothing here needs a decision — it is so you know how much weight to put on the document."
ws["A1"].font = Font(name=F, size=10, italic=True, color="666666")
ws.merge_cells("A1:D1")
hdr = ["Verdict", "What the analysis says", "What I found", "So"]
for i, w in enumerate([16, 52, 66, 40], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for i, h in enumerate(hdr, start=1):
    ws.cell(row=2, column=i, value=h)
style_header(ws, 2, len(hdr))

CHECKS = [
    ("Confirmed", "82 priorities quoted across the issues", "Every one matches QACR-APP-FR-01 Rev 1.6 exactly. Zero discrepancies.", "The transcription can be relied on, as it claims."),
    ("Confirmed", "Five identifiers appear nowhere: FR-IMG-018, FR-IMG-019, FR-SHR-009, FR-SHR-010, FR-TIM-013", "Genuinely absent from the 184 in scope, the 19 Appendix D rows and all 18 \"formerly FR-\" notes in the backlog. Every other number in every one of the 23 prefixes is accounted for.", "Real. DC-08."),
    ("Confirmed", "Every stale note it quotes", "FR-AUT-009's stray semicolon, FR-PRT-004 naming a withdrawn identifier, the three \"priority 0\" rows, BL-13 still deferred, the FR-CFG-005 over-claim — all present as quoted.", "All nine are on the Doc corrections tab."),
    ("Needs your input", "M1 is mid-October 2026, M2 early-December, M3 and M4 both early-January 2027; M3 and M4 share a date", "No date appears anywhere in our requirements document. The calendar is the analysis's own, from somewhere I cannot see.", "Confirm or correct it. The must-have/nice-to-have rule holds either way, but the urgency framing of half the issues depends on it."),
    ("Overstated", "\"FR-ALG-008 and FR-AUT-009 appear in the exclusions appendix and in the in-scope list at the same time\"", "The Appendix D rows are titled \"FR-ALG-008 (earlier wording)\" and \"FR-AUT-009 (earlier wording)\", so they are distinguishable, not duplicated. The real defect is that the appendix mixes three kinds of row without saying so.", "Reduced to a presentation fix, DC-07, rather than a contradiction."),
    ("Overstated", "\"If M4 is not delivered the must-have portal ships with no access protection at all\"", "The backend enforces the PIN registration check, the five-attempt lockout and the 24-hour setup window server-side, per SRS-BE PIN.1 to PIN.7 cited in those same requirements.", "The inversion is real — CR-03 — but the exposure is the application-side gate, not the whole control."),
    ("Overstated", "\"FR-TXT-021 and FR-ALG-012 each name the other as the dependency, so the document does not settle which comes first\"", "They are cross-references, not a circular dependency. FR-ALG-012 supplies the reason category and FR-TXT-021 consumes it; at priorities 2 then 3 the order is already correct.", "The defect is only that FR-RES-005 at priority 1 consumes both. Folded into CR-07."),
    ("Incomplete", "\"FR-LCM-008 is nice-to-have regression testing of the timing and test-state controls\"", "True, but FR-LCM-007 — each timer and window verified separately, including under backgrounding — is priority 3, so timing verification is not wholly nice-to-have.", "CR-09 stands but is narrower than stated."),
    ("Bookkeeping", "\"20 new issues, DT15-01 to DT15-20\"", "DT15-20 is numbered twice, with two overlapping write-ups of the same RA 4.17 point. There are 19 distinct issues. The task preamble also says \"read all 16 issues\".", "This worksheet has 19 of theirs plus 2 of ours."),
    ("We found, it did not", "—", "Test data upload is priority 1 while TLS enforcement and certificate pinning are priority 3, so a demonstration build transmits over whatever the platform negotiates.", "CR-21."),
    ("Both found", "Seven of the inversions", "CR-01, CR-02, CR-03, CR-04, CR-07, CR-10 and the FR-CAM-002 half of CR-12 are all in section 4 of QACR-APP-EPIC-01, reached independently by grouping requirements into features.", "Two methods, same conclusions. Reasonable confidence in all seven."),
]
r = 3
VCOL = {"Confirmed": "1E6B3A", "Overstated": "8A6D3B", "Incomplete": "8A6D3B",
        "Needs your input": "9C1F1F", "Bookkeeping": "666666",
        "We found, it did not": "1F3864", "Both found": "1F3864"}
for c in CHECKS:
    put(ws, r, list(c), alt=(r % 2 == 0))
    ws.cell(row=r, column=1).font = Font(name=F, size=9, bold=True, color=VCOL[c[0]])
    ws.row_dimensions[r].height = 56
    r += 1
ws.freeze_panes = "A3"

for ws in wb:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUT)
print("wrote", OUT)
print("conflicts:", len(CONFLICTS), "corrections:", len(CORR), "checks:", len(CHECKS))
